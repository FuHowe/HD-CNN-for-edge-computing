import os
import subprocess
import mypath
import openpyxl
import myfile
import numpy as np
from openpyxl.styles import Font, colors, Alignment


def lw_download_data(file_name):
    print('������STM32��������...')
    current_path = os.getcwd()
    
    
    
    



    
    
    
    
    
    
    
    
    os.chdir(mypath.programmer_path) 
    file_name = os.path.join(current_path, mypath.model_path.lstrip('./'), file_name)
    input_str = 'cmd.exe /c STM32_Programmer_CLI.exe -c port=SWD  mode=NORMAL -d ' + file_name + ' 0x08020000 -v -s '
    out = subprocess.Popen(input_str, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    
    os.chdir(current_path) 
    out.wait() 


    if (out.returncode == 0):
        print('�����������\n')
    else:
        print('[Error]����STM32��������ʧ�ܣ����ST-Link���ӡ��ļ�·����exe��bin�Ƿ����...')

        
        for line in out.stdout.readlines():
            print(line)
    



def dt_export_result(count, best_correct_rate, \
                     image_size, input_dim_1, output_dim_1, kernel_1, pooling_1, input_dim_2, output_dim_2, kernel_2, pooling_2, input_dim_3, output_dim_3, kernel_3, pooling_3, full_connect_len, classification_out,\
                     head_len,conv_1, conv_2, conv_3, linear_1, img_size, total_size, max_ram):
    
    excel_name = 'Result.xlsx'
    sheet_name = 'Parameter'
        
    
    if not os.path.exists(excel_name):        
        excel = openpyxl.Workbook() 
        excel.remove(excel['Sheet']) 
        performance_sheet = excel.create_sheet(sheet_name) 

        
        sheet_title = ['���', '��߾���%', 'ͼ��ߴ�', 
                       '��1������ά��', '��1�����ά��', '��1������','��1��ػ���', 
                       '��2������ά��', '��2�����ά��', '��2������','��2��ػ���', 
                       '��3������ά��', '��3�����ά��', '��3������','��3��ػ���', 
                       'ȫ�������볤��', '�����������', 
                       '�ļ�ͷ��(KB)', 
                       '��1��Ȩ�ش�С(KB)', '��2��Ȩ�ش�С(KB)', '��3��Ȩ�ش�С(KB)', 
                       '���Բ�Ȩ�ش�С(KB)','ͼ���С(KB)', 'Flash�ܴ�С(KB)', 'Conv1_RAM','Conv2_RAM','Conv3_RAM','MAX_RAM']
        
        
        for i in range(0, len(sheet_title)):
            performance_sheet.cell(1, i + 1).value = sheet_title[i]
        
        excel.save(excel_name)
        excel.close()
    
    
    exist_excel = openpyxl.load_workbook(excel_name) 
    exist_sheet = exist_excel[sheet_name] 
    
    
    write_data = [count, best_correct_rate, image_size, 
                  input_dim_1, output_dim_1, kernel_1, pooling_1, 
                  input_dim_2, output_dim_2, kernel_2, pooling_2, 
                  input_dim_3, output_dim_3, kernel_3, pooling_3, 
                  full_connect_len, classification_out,
                  head_len,
                  conv_1, conv_2, conv_3, 
                  linear_1, img_size, total_size,max_ram[0],max_ram[1],max_ram[2],max_ram[3]]

    exist_sheet.append(write_data) 

    exist_excel.save(excel_name)
    exist_excel.close()



def lw_delete_file(folder_path, appendix):
    for file_name in os.listdir(folder_path):
       file_path = os.path.join(folder_path, file_name)
       if file_name.endswith(appendix):
           
           os.remove(file_path)


        
        
        
        
        
        
        
        

    
